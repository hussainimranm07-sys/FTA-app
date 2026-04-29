"""
FTA Risk Allocator v11
=======================
Full DAG (Directed Acyclic Graph) support — nodes can have MULTIPLE parents.
This correctly models shared events (SF-05a, SF-05b) appearing in multiple
AND-gate fault tree branches.

KEY ARCHITECTURE CHANGES vs v10:
  - nodes[id]["parents"] = LIST of parent IDs  (was single "parent")
  - edges stored separately: edges = [ {from, to, gate} ]
    gate on the EDGE = gate type of the parent node at that connection
  - allocate()  : DAG traversal, node gets MIN allocation across all paths
  - rollup()    : DAG traversal, node contributes its value to each parent path
  - shared sync : same label → same achieved value (worst-case)

GATE SEMANTICS (on parent node, applied to ALL its children via that parent):
  OR  gate parent: child_T = parent_T / n_children
  AND gate parent: child_T = parent_T ^ (1/n_children)

ROLLUP:
  OR  gate parent: parent_A = sum(children_A)
  AND gate parent: parent_A = product(children_A)

MULTI-PATH ALLOCATION:
  A node with multiple parents receives a budget from EACH parent path.
  Its effective budget = MIN of all paths (most conservative / strictest requirement).
"""

import streamlit as st
import streamlit.components.v1 as components
import json, math, io, datetime, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="FTA Allocator v11", page_icon="⚛", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600;700&family=DM+Sans:wght@300;400;500;600&display=swap');
*{box-sizing:border-box}
html,body,[class*="css"]{font-family:'DM Sans',sans-serif}
.stApp{background:#080c14;color:#d4dde8}
section[data-testid="stSidebar"]{background:#0c1220!important;border-right:1px solid #1e2d45}
section[data-testid="stSidebar"] *{color:#d4dde8!important}
.stButton>button{background:#0f1d30!important;border:1px solid #1e3a5f!important;
  color:#7ab8e8!important;border-radius:5px!important;font-family:'DM Sans',sans-serif!important;
  font-size:0.8rem!important;transition:all .15s!important}
.stButton>button:hover{background:#1e3a5f!important;color:#b8d8f5!important;border-color:#4a8cc2!important}
.stTabs [data-baseweb="tab-list"]{background:#0c1220;border-bottom:1px solid #1e2d45}
.stTabs [data-baseweb="tab"]{color:#5a7a9a!important;font-family:'DM Sans',sans-serif}
.stTabs [aria-selected="true"]{color:#7ab8e8!important;border-bottom:2px solid #4a8cc2!important}
div[data-testid="stExpander"]{background:#0c1220;border:1px solid #1e2d45;border-radius:6px}
.stTextInput input,.stNumberInput input,.stSelectbox select,.stTextArea textarea{
  background:#0c1220!important;border:1px solid #1e2d45!important;color:#d4dde8!important;
  font-family:'JetBrains Mono',monospace!important;border-radius:5px!important}
.stSelectbox [data-baseweb="select"]>div{background:#0c1220!important;border-color:#1e2d45!important}
hr{border-color:#1e2d45!important}
.vbadge{font-family:'JetBrains Mono',monospace;font-size:0.78rem;font-weight:600;
  display:inline-block;padding:2px 9px;border-radius:4px;letter-spacing:.5px}
.vb-alloc{background:#0d2036;border:1px solid #1e4a7a;color:#5aabff}
.vb-ach{background:#0d1f18;border:1px solid #1e4a32;color:#4ade80}
.vb-ach.over{background:#200d0d;border-color:#4a1e1e;color:#f87171}
.vb-none{background:#141820;border:1px solid #2a3040;color:#4a5a6a}
.nb{display:inline-block;padding:1px 7px;border-radius:3px;font-size:0.67rem;font-weight:700;
  font-family:'JetBrains Mono',monospace;letter-spacing:.5px}
.nb-HZ{background:#2d1200;color:#fb923c;border:1px solid #7c3300}
.nb-SF{background:#001830;color:#60a5fa;border:1px solid #1e4878}
.nb-FF{background:#001a10;color:#34d399;border:1px solid #065f35}
.nb-IF{background:#1a0030;color:#c084fc;border:1px solid #5b21b6}
.nb-AND{background:#200025;color:#e879f9;border:1px solid #7e22ce}
.st-shared{display:inline-block;padding:1px 6px;border-radius:3px;font-size:0.62rem;
  font-weight:700;background:#1a1000;color:#fbbf24;border:1px solid #78350f;margin-left:4px}
.st-multi{display:inline-block;padding:1px 6px;border-radius:3px;font-size:0.62rem;
  font-weight:700;background:#001828;color:#38bdf8;border:1px solid #0369a1;margin-left:4px}
.callout{background:#0c1628;border:1px solid #1e3a5f;border-left:3px solid #4a8cc2;
  border-radius:6px;padding:10px 14px;font-size:0.8rem;color:#7ab8e8;line-height:1.6;margin:8px 0}
.callout.warn{border-left-color:#f59e0b;color:#fcd34d;background:#100e04}
.callout.ok{border-left-color:#22c55e;color:#86efac;background:#04100a}
.app-header{background:linear-gradient(135deg,#0c1e38 0%,#080c14 100%);
  border:1px solid #1e3a5f;border-left:4px solid #4a8cc2;border-radius:8px;
  padding:16px 24px;margin-bottom:18px}
.app-header h1{font-family:'JetBrains Mono',monospace;font-size:1.3rem;
  color:#7ab8e8;margin:0 0 3px;letter-spacing:-0.5px}
.app-header p{color:#4a6a8a;margin:0;font-size:0.75rem}
.save-ok{background:#081208;border:1px solid #1e4032;border-radius:5px;
  padding:5px 12px;font-size:0.72rem;color:#4ade80;display:inline-block}
.save-no{background:#100e04;border:1px solid #2a2810;border-radius:5px;
  padding:5px 12px;font-size:0.72rem;color:#a3a380;display:inline-block}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════
TYPE_STR  = {"HZ":"#fb923c","SF":"#60a5fa","FF":"#34d399","IF":"#c084fc","AND":"#e879f9"}
SAVE_FILE = "fta_save_v11.json"

# Valid parent types for each node type
VALID_PARENT_TYPES = {
    "SF":  ["HZ","SF","AND"],
    "FF":  ["SF","FF","AND"],
    "IF":  ["FF","SF"],
    "AND": ["HZ","SF","FF"],
}

# ══════════════════════════════════════════════════════════════════
# DEFAULT DATA — HZ04 Pressurized Fluid Hazard (from your FT50–FT61)
# ══════════════════════════════════════════════════════════════════
def build_default_data():
    """
    Build the complete HZ04 fault tree from your PDFs.

    Structure:
    HZ04 (OR, target=1e-8)
    ├── SF-17  (OR gate — any of the FT branches causes SF-17)
    │   ├── AND-FT50  (AND: FF-90 AND thermal runaway)
    │   │   ├── FF-90  (OR: 10 IF causes)
    │   │   │   ├── IF-200..IF-301 (10 initiating failures)
    │   │   └── OR-TR1 (OR: SF-05a OR SF-05b)  ← SF-05a,SF-05b shared!
    │   │       ├── SF-05a (shared)
    │   │       └── SF-05b (shared)
    │   ├── AND-FT52  (AND: FF-52 AND thermal runaway)
    │   │   ├── FF-52  (OR: 3 IF causes)
    │   │   └── OR-TR2 → SF-05a (shared), SF-05b (shared)
    │   ├── AND-FT53  (AND: FF-53 AND thermal runaway)
    │   │   ├── FF-53  (OR: 3 IF causes)
    │   │   └── OR-TR3 → SF-05a (shared), SF-05b (shared)
    │   ├── AND-FT61  (AND: FF-91 AND SF-04c)
    │   │   ├── FF-91 (1 IF)
    │   │   └── SF-04c
    │   └── FF-74  (OR: ~20 IF causes — uncontrolled depressurization)
    └── SF-17b (placeholder — to be expanded)
    """
    nodes = {}
    edges = []  # {from_id, to_id, from_gate}  — gate on the FROM (parent) side
    hz_targets = {}
    nxt = [200]  # use list so inner func can mutate

    def nid():
        v = f"N{nxt[0]:04d}"; nxt[0] += 1; return v

    def add_node(id_, label, name, ntype, gate="OR", achieved=None, desc=""):
        nodes[id_] = {"id": id_, "label": label, "name": name,
                      "type": ntype, "gate": gate,
                      "achieved": achieved, "desc": desc, "parents": []}

    def add_edge(from_id, to_id):
        # gate used is the FROM node's gate (how it combines its children)
        edges.append({"from": from_id, "to": to_id})
        if to_id not in nodes[from_id].get("_children_ids", []):
            nodes[from_id].setdefault("_children_ids", []).append(to_id)
        nodes[to_id]["parents"].append(from_id)

    # ── HZ04 root ────────────────────────────────────────────────
    add_node("HZ04", "HZ04", "Pressurized Fluid Hazard", "HZ", gate="OR")
    hz_targets["HZ04"] = 1e-8

    # ── SF-17 and SF-17b ─────────────────────────────────────────
    add_node("SF17",  "SF-17",  "Sudden pressure release through an unintended path", "SF", gate="OR")
    add_node("SF17b", "SF-17b", "Sudden pressure release through a second path",       "SF", gate="OR")
    add_edge("HZ04", "SF17")
    add_edge("HZ04", "SF17b")

    # ── Shared thermal runaway events ────────────────────────────
    # SF-05a and SF-05b will appear under multiple OR groups
    add_node("SF05a", "SF-05a", "Single cell Thermal Runaway",                "SF", gate="OR")
    add_node("SF05b", "SF-05b", "Multiple cell Thermal Runaway (not by 05a)", "SF", gate="OR")

    # ── FT50: AND gate — FF-90 AND (SF-05a OR SF-05b) ────────────
    and50 = nid()
    add_node(and50, "AND-FT50", "Combined fault — valve stuck closed AND thermal runaway", "AND", gate="AND")
    add_edge("SF17", and50)

    add_node("FF90", "FF-90", "Degassing valves stuck in the closed position", "FF", gate="OR")
    add_edge(and50, "FF90")

    or_tr50 = nid()
    add_node(or_tr50, "OR-TR50", "Thermal runaway (FT50 branch)", "SF", gate="OR")
    add_edge(and50, or_tr50)
    add_edge(or_tr50, "SF05a")
    add_edge(or_tr50, "SF05b")

    # FF-90 initiating failures (10 causes from FT50 PDF)
    ff90_ifs = [
        ("IF200", "IF-200", "Degassing valves mechanical damage due to shock & vibration during operation"),
        ("IF201", "IF-201", "Degassing valves mechanical damage due to crash"),
        ("IF202", "IF-202", "Degassing valves mechanical damage during manufacturing or logistics"),
        ("IF203", "IF-203", "Degassing valves manufacturing errors"),
        ("IF206", "IF-206", "Degassing valves material degradation"),
        ("IF207", "IF-207", "Pressure built is not enough for degassing valves to open"),
        ("IF090", "IF-090", "Degassing valves corrosion"),
        ("IF378", "IF-378", "Degassing valves contamination of particles during manufacturing"),
        ("IF244", "IF-244", "FOAM blocking degassing channel"),
        ("IF301", "IF-301", "Degassing channel contamination with foreign material (Incl Mud)"),
    ]
    for iid, lbl, nm in ff90_ifs:
        add_node(iid, lbl, nm, "IF")
        add_edge("FF90", iid)

    # ── FT52: AND gate — FF-52 AND (SF-05a OR SF-05b) ────────────
    and52 = nid()
    add_node(and52, "AND-FT52", "Combined fault — cell disk not opened AND thermal runaway", "AND", gate="AND")
    add_edge("SF17", and52)

    add_node("FF52", "FF-52", "Cell disk not opened as intended", "FF", gate="OR")
    add_edge(and52, "FF52")

    or_tr52 = nid()
    add_node(or_tr52, "OR-TR52", "Thermal runaway (FT52 branch)", "SF", gate="OR")
    add_edge(and52, or_tr52)
    add_edge(or_tr52, "SF05a")   # ← shared node gets second parent
    add_edge(or_tr52, "SF05b")   # ← shared node gets second parent

    ff52_ifs = [
        ("IF310", "IF-310", "Cell support mechanical damage due to manufacturing / assembly"),
        ("IF008", "IF-008", "Contamination of particles during cell manufacturing"),
        ("IF319", "IF-319", "FOAM build-up into cell support venting"),
    ]
    for iid, lbl, nm in ff52_ifs:
        add_node(iid, lbl, nm, "IF")
        add_edge("FF52", iid)

    # ── FT53: AND gate — FF-53 AND (SF-05a OR SF-05b) ────────────
    and53 = nid()
    add_node(and53, "AND-FT53", "Combined fault — degassing valves stuck open AND thermal runaway", "AND", gate="AND")
    add_edge("SF17", and53)

    add_node("FF53", "FF-53", "Degassing valves stuck in open position", "FF", gate="OR")
    add_edge(and53, "FF53")

    or_tr53 = nid()
    add_node(or_tr53, "OR-TR53", "Thermal runaway (FT53 branch)", "SF", gate="OR")
    add_edge(and53, or_tr53)
    add_edge(or_tr53, "SF05a")   # ← shared node gets third parent
    add_edge(or_tr53, "SF05b")   # ← shared node gets third parent

    ff53_ifs = [
        ("IF317", "IF-317", "Contamination of particles during manufacturing (EVD)"),
        ("IF320", "IF-320", "FOAM intrusion into degassing channels"),
        ("IF378b","IF-378b","Degassing valves contamination of particles during manufacturing"),
    ]
    for iid, lbl, nm in ff53_ifs:
        add_node(iid, lbl, nm, "IF")
        add_edge("FF53", iid)

    # ── FT61: AND gate — FF-91 (dust/particle blockage) AND SF-04c ──
    and61 = nid()
    add_node(and61, "AND-FT61", "Combined fault — degassing blockage AND over-temperature PDU", "AND", gate="AND")
    add_edge("SF17", and61)

    add_node("FF91", "FF-91",  "Degassing blockage due to dust or particles", "FF", gate="OR")
    add_node("SF04c","SF-04c", "PDU over-temperature (conduction, convection, or radiation)", "SF", gate="OR")
    add_node("IF289","IF-289", "Pressure increased inside PDU due to breather blockage", "IF")
    add_edge(and61, "FF91")
    add_edge(and61, "SF04c")
    add_edge("FF91",  "IF289")

    # ── FT51: FF-74 uncontrolled depressurization (~20 IFs) ──────
    add_node("FF74", "FF-74", "Uncontrolled depressurisation / disconnection", "FF", gate="OR")
    add_edge("SF17", "FF74")

    ft51_ifs = [
        ("IF251", "IF-251", "Gateway valves material fatigue"),
        ("IF394", "IF-394", "Boundary valves mechanical damage due to shock & vibration during operation"),
        ("IF009", "IF-009", "Gateway valves mechanical damage due to crash"),
        ("IF985", "IF-985", "Gateway valves mechanical damage due to manufacturing logistics and transport"),
        ("IF985b","IF-985b","Gateway valves manufacturing errors"),
        ("IF370", "IF-370", "Mass management in fluid system"),
        ("IF241", "IF-241", "Flammability due to internal cooling media"),
        ("IF849", "IF-849", "ADI PRO disconnection error / Manufacturing error"),
        ("IF168", "IF-168", "Thin wall component foundation/mounting and drive train failure"),
        ("IF168b","IF-168b","Thin wall component foundation/mounting and due to crash"),
        ("IF221", "IF-221", "PDA miscalculation plus corrective operational handling in Electrical Motor"),
        ("IF460", "IF-460", "Pop-up mechanical design or manufacturing errors (intermediate)"),
        ("IF525", "IF-525", "POL friction counting material failure"),
        ("IF526", "IF-526", "POL friction counting material failure"),
        ("IF886", "IF-886", "Cleaning / setup environmental conditions during manufacturing"),
        ("IF003", "IF-003", "Weld / bond breaks in the design sustained cuts due to repeated loads"),
        ("IF500", "IF-500", "Gaskets/seals internal leaks become coolant Environment"),
        ("IF501", "IF-501", "Absolute Coolants disconnect from driving"),
        ("IF502", "IF-502", "Propulsion tube compartment malfunction"),
        ("IF503", "IF-503", "Pop-up mechanical design or manufacturing errors at connections"),
    ]
    for iid, lbl, nm in ft51_ifs:
        add_node(iid, lbl, nm, "IF")
        add_edge("FF74", iid)

    # ── SF-17b placeholder (to be expanded by user) ──────────────
    # SF-17b gets an OR-group with a placeholder FF
    add_node("FF17b1","FF-17b-1","Second path failure mode (expand as needed)", "FF", gate="OR")
    add_edge("SF17b", "FF17b1")
    add_node("IF17b1","IF-17b-1","Initiating failure for SF-17b (placeholder)", "IF")
    add_edge("FF17b1","IF17b1")

    # Clean up internal helper key
    for n in nodes.values():
        n.pop("_children_ids", None)

    return nodes, edges, hz_targets, nxt[0]


# ══════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════
def _init():
    if "nodes" not in st.session_state:
        st.session_state.nodes = {}
    if "edges" not in st.session_state:
        st.session_state.edges = []
    if "hz_targets" not in st.session_state:
        st.session_state.hz_targets = {}
    if "nxt" not in st.session_state:
        st.session_state.nxt = 1
    if "loaded_default" not in st.session_state:
        st.session_state.loaded_default = False

_init()


# ══════════════════════════════════════════════════════════════════
# GRAPH HELPERS
# ══════════════════════════════════════════════════════════════════
def get_children(edges, pid):
    """Return list of child node IDs for a given parent."""
    return [e["to"] for e in edges if e["from"] == pid]

def get_parents(edges, nid):
    """Return list of parent node IDs for a given node."""
    return [e["from"] for e in edges if e["to"] == nid]

def hz_roots(nodes):
    return [n for n in nodes.values() if n["type"] == "HZ"]

def all_node_ids_reachable(nodes, edges):
    """BFS from all HZ roots, returns ordered list of node IDs."""
    roots = [n["id"] for n in hz_roots(nodes)]
    out, queue, seen = [], list(roots), set(roots)
    while queue:
        nid = queue.pop(0)
        if nid not in nodes: continue
        out.append(nid)
        for cid in get_children(edges, nid):
            if cid not in seen and cid in nodes:
                seen.add(cid); queue.append(cid)
    # add any disconnected nodes
    for nid in nodes:
        if nid not in seen:
            out.append(nid)
    return out

def nodes_with_label(nodes, label):
    if not label: return []
    return [nid for nid, n in nodes.items() if n.get("label","") == label]

def fmt(v, dash="–"):
    if v is None: return dash
    if v == 0:   return "0.000E+00"
    return f"{v:.3E}"

def next_id():
    nid = f"N{st.session_state.nxt:04d}"
    st.session_state.nxt += 1
    return nid

def depth_in_dag(nodes, edges, nid):
    """Longest path from any HZ root to nid."""
    memo = {}
    def _d(n):
        if n in memo: return memo[n]
        pars = get_parents(edges, n)
        if not pars:
            memo[n] = 0; return 0
        d = 1 + max(_d(p) for p in pars if p in nodes)
        memo[n] = d; return d
    return _d(nid)


# ══════════════════════════════════════════════════════════════════
# CORE ENGINE 1: DAG ALLOCATION (top-down)
# Each node receives budgets from ALL its parents.
# Its effective budget = MIN across all incoming paths.
# ══════════════════════════════════════════════════════════════════
def allocate(nodes, edges, hz_targets):
    """
    Top-down DAG allocation.
    Returns {nid: float}  — effective budget for each node.

    Algorithm:
    1. Topological sort (Kahn's algorithm from HZ roots)
    2. For each node in topo order, compute what budget it gives each child
    3. Each child collects ALL incoming budgets → takes the MIN
    """
    # Kahn's topological sort
    in_deg = {nid: 0 for nid in nodes}
    for e in edges:
        if e["to"] in in_deg:
            in_deg[e["to"]] += 1

    queue = [nid for nid, d in in_deg.items() if d == 0 and nodes[nid]["type"] == "HZ"]
    topo = []
    remaining = dict(in_deg)
    visited = set()
    q = list(queue)
    while q:
        nid = q.pop(0)
        if nid in visited: continue
        visited.add(nid); topo.append(nid)
        for cid in get_children(edges, nid):
            if cid in remaining:
                remaining[cid] -= 1
                if remaining[cid] == 0:
                    q.append(cid)

    # Budget propagation
    budget = {}  # nid -> effective budget (min of all incoming)
    incoming = {}  # nid -> list of budgets offered by parents

    for hz in hz_roots(nodes):
        budget[hz["id"]] = hz_targets.get(hz["id"], 1e-7)

    for nid in topo:
        if nid not in budget: continue
        b = budget[nid]
        kids = get_children(edges, nid)
        if not kids: continue
        n_kids = len(kids)
        gate = nodes[nid].get("gate", "OR")
        for kid in kids:
            if gate == "AND":
                child_b = (b ** (1.0 / n_kids)) if b > 0 else 0.0
            else:
                child_b = b / n_kids if n_kids > 0 else 0.0
            incoming.setdefault(kid, []).append(child_b)

    # Each node's effective budget = min of all incoming offers
    result = {}
    for hz in hz_roots(nodes):
        result[hz["id"]] = budget[hz["id"]]

    for nid in nodes:
        if nid in incoming and incoming[nid]:
            result[nid] = min(incoming[nid])
        elif nodes[nid]["type"] == "HZ":
            result[nid] = hz_targets.get(nid, 1e-7)

    return result


# ══════════════════════════════════════════════════════════════════
# CORE ENGINE 2: DAG ROLLUP (bottom-up achieved values)
# ══════════════════════════════════════════════════════════════════
def rollup(nodes, edges):
    """
    Bottom-up DAG rollup of achieved values.
    Shared nodes (multiple parents): their achieved value is computed once
    and contributes to EACH parent path independently.
    Returns {nid: float|None}
    """
    cache = {}

    def _compute(nid, visiting=None):
        if visiting is None: visiting = set()
        if nid in cache: return cache[nid]
        if nid in visiting:  # cycle guard
            return None
        visiting = visiting | {nid}

        n = nodes.get(nid)
        if n is None:
            cache[nid] = None; return None

        kids = get_children(edges, nid)

        if not kids:
            # Leaf: use manual value
            cache[nid] = n.get("achieved")
            return cache[nid]

        # Manual override on non-leaf takes precedence
        if n.get("achieved") is not None:
            cache[nid] = n["achieved"]
            return cache[nid]

        child_vals = [_compute(k, visiting) for k in kids]

        if any(v is None for v in child_vals):
            cache[nid] = None; return None

        gate = n.get("gate", "OR")
        if gate == "AND":
            val = 1.0
            for v in child_vals: val *= v
        else:
            val = sum(child_vals)

        cache[nid] = val
        return val

    for hz in hz_roots(nodes):
        _compute(hz["id"])
    for nid in nodes:
        if nid not in cache:
            _compute(nid)

    return cache


# ══════════════════════════════════════════════════════════════════
# CORE ENGINE 3: SHARED FAILURE SYNC
# ══════════════════════════════════════════════════════════════════
def sync_shared(nodes, changed_nid, new_value):
    label = nodes[changed_nid].get("label","")
    peers = nodes_with_label(nodes, label)
    if len(peers) <= 1:
        old = nodes[changed_nid].get("achieved")
        nodes[changed_nid]["achieved"] = new_value
        return [(changed_nid, old, new_value)] if old != new_value else []
    existing = [nodes[p].get("achieved") for p in peers if nodes[p].get("achieved") is not None]
    worst = max([new_value] + existing) if existing else new_value
    log = []
    for pid in peers:
        old = nodes[pid].get("achieved")
        if old != worst:
            nodes[pid]["achieved"] = worst
            log.append((pid, old, worst))
    return log


# ══════════════════════════════════════════════════════════════════
# PERSISTENCE
# ══════════════════════════════════════════════════════════════════
def state_to_dict():
    return {
        "version": "v11",
        "saved_at": datetime.datetime.utcnow().isoformat() + "Z",
        "nodes": st.session_state.nodes,
        "edges": st.session_state.edges,
        "hz_targets": st.session_state.hz_targets,
        "nxt": st.session_state.nxt,
    }

def dict_to_state(d):
    st.session_state.nodes      = d.get("nodes", {})
    st.session_state.edges      = d.get("edges", [])
    st.session_state.hz_targets = d.get("hz_targets", {})
    st.session_state.nxt        = d.get("nxt", 1)

def save_to_file():
    try:
        with open(SAVE_FILE, "w") as f:
            json.dump(state_to_dict(), f, indent=2)
        return True
    except Exception:
        return False

def load_from_file():
    try:
        if os.path.exists(SAVE_FILE):
            with open(SAVE_FILE) as f:
                d = json.load(f)
            dict_to_state(d)
            return d.get("saved_at","")
    except Exception:
        pass
    return None

# ── Boot sequence ──────────────────────────────────────────────
if "auto_loaded" not in st.session_state:
    st.session_state.auto_loaded = True
    ts = load_from_file()
    if ts:
        st.session_state["_last_saved"] = ts
    elif not st.session_state.nodes:
        # No saved file → load the HZ04 default data
        nd, ed, ht, nxt_val = build_default_data()
        st.session_state.nodes      = nd
        st.session_state.edges      = ed
        st.session_state.hz_targets = ht
        st.session_state.nxt        = nxt_val
        st.session_state.loaded_default = True

if st.session_state.nodes:
    save_to_file()
    st.session_state["_last_saved"] = datetime.datetime.utcnow().strftime("%H:%M:%S UTC")


# ══════════════════════════════════════════════════════════════════
# VISUALIZATION (DAG-aware canvas)
# ══════════════════════════════════════════════════════════════════
def build_canvas(nodes, edges, alloc, rolled):
    order = all_node_ids_reachable(nodes, edges)
    hz_ids = [n["id"] for n in hz_roots(nodes)]
    palette = ["#4a8cc2","#22c55e","#f59e0b","#e879f9","#38bdf8","#f87171","#a3e635","#06b6d4"]

    # Colour by HZ ancestry (first parent HZ wins for colour)
    def hz_anc(nid, seen=None):
        if seen is None: seen = set()
        if nid in seen: return None
        seen.add(nid)
        n = nodes.get(nid)
        if not n: return None
        if n["type"] == "HZ": return nid
        pars = get_parents(edges, nid)
        for p in pars:
            r = hz_anc(p, seen)
            if r: return r
        return None

    hz_color = {hid: palette[i % len(palette)] for i, hid in enumerate(hz_ids)}

    node_data, edge_data = [], []
    for nid in order:
        if nid not in nodes: continue
        n = nodes[nid]
        t_val = alloc.get(nid)
        a_val = rolled.get(nid)
        status = ("pass" if (t_val and a_val and a_val <= t_val) else
                  "fail" if (t_val and a_val and a_val > t_val) else "na")
        ha = hz_anc(nid)
        pars = get_parents(edges, nid)
        peers = nodes_with_label(nodes, n.get("label",""))
        node_data.append({
            "id": nid, "label": n.get("label", nid), "name": n.get("name",""),
            "type": n["type"], "gate": n.get("gate","–"),
            "T": fmt(t_val), "A": fmt(a_val),
            "status": status, "color": hz_color.get(ha, "#4a8cc2"),
            "multi_parent": len(pars) > 1,
            "shared": len(peers) > 1,
        })

    for e in edges:
        if e["from"] in nodes and e["to"] in nodes:
            gate = nodes[e["from"]].get("gate","OR")
            edge_data.append({"from": e["from"], "to": e["to"], "gate": gate})

    nj = json.dumps(node_data)
    ej = json.dumps(edge_data)

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#080c14;font-family:'JetBrains Mono',monospace;overflow:hidden;user-select:none}}
#cv{{cursor:grab;display:block}}#cv.pan{{cursor:grabbing}}
#hud{{position:absolute;top:10px;left:10px;display:flex;flex-direction:column;gap:4px;z-index:10}}
.hbtn{{background:#0c1628;border:1px solid #1e3a5f;color:#7ab8e8;padding:4px 10px;border-radius:4px;
  cursor:pointer;font-size:10px;font-family:inherit;white-space:nowrap}}
.hbtn:hover{{background:#1e3a5f}}
#srch{{position:absolute;top:10px;left:50%;transform:translateX(-50%);
  background:#0c1628;border:1px solid #1e3a5f;color:#d4dde8;padding:5px 12px;
  border-radius:4px;font-size:11px;font-family:inherit;width:220px;outline:none}}
#srch:focus{{border-color:#4a8cc2}}
#tip{{position:absolute;background:#0c1e38;border:1px solid #1e3a5f;border-radius:6px;
  padding:8px 12px;font-size:10px;color:#d4dde8;z-index:20;display:none;
  max-width:280px;pointer-events:none;line-height:1.6;box-shadow:0 8px 24px rgba(0,0,0,.9)}}
#leg{{position:absolute;bottom:8px;left:10px;font-size:9px;color:#3a5a7a;display:flex;gap:10px;flex-wrap:wrap}}
.ld{{display:inline-block;width:7px;height:7px;border-radius:50%;margin-right:3px;vertical-align:middle}}
</style></head><body>
<canvas id="cv"></canvas>
<div id="hud">
  <button class="hbtn" onclick="layout()">⊞ Layout</button>
  <button class="hbtn" onclick="zoomIn()">＋ Zoom</button>
  <button class="hbtn" onclick="zoomOut()">－ Zoom</button>
  <button class="hbtn" onclick="resetView()">⌖ Reset</button>
  <button class="hbtn" id="simBtn" onclick="toggleSim()">⟳ Physics OFF</button>
</div>
<input id="srch" type="text" placeholder="Search label / name…" oninput="onSearch(this.value)">
<div id="tip"></div>
<div id="leg">
  <span><span class="ld" style="background:#22c55e"></span>A≤T (OK)</span>
  <span><span class="ld" style="background:#f87171"></span>A>T (Exceeds)</span>
  <span><span class="ld" style="background:#3a5a7a"></span>No data</span>
  <span><span class="ld" style="background:#fbbf24;border-radius:2px"></span>Shared failure</span>
  <span><span class="ld" style="background:#38bdf8;border-radius:2px"></span>Multi-parent (DAG)</span>
  <span style="color:#2a4a6a">drag·scroll=zoom·hover=info</span>
</div>
<script>
const NODES={nj};
const EDGES={ej};
const BW=165,BH=66,GR=11;
const SCOL={{pass:"#22c55e",fail:"#f87171",na:"#2a3a4a"}};
const TFIL={{HZ:"#1a0e00",SF:"#00101e",FF:"#00110a",IF:"#0e0018",AND:"#130018"}};
const TSTR={{HZ:"#fb923c",SF:"#60a5fa",FF:"#34d399",IF:"#c084fc",AND:"#e879f9"}};
const cv=document.getElementById('cv');
const ctx=cv.getContext('2d');
let scale=1,panX=0,panY=60,sim=false;
let drag=null,dragOff={{x:0,y:0}},isPan=false,lastP={{x:0,y:0}};
let pos={{}},searchHL=new Set(),collapsed=new Set();

function resize(){{cv.width=window.innerWidth;cv.height=window.innerHeight;}}
window.addEventListener('resize',()=>{{resize();draw();}});
resize();

function layout(){{
  pos={{}};
  // Compute depth for each node (longest path from root)
  const depth={{}};
  function getDepth(id,vis){{
    if(id in depth)return depth[id];
    if(vis.has(id))return 0;
    vis=new Set(vis);vis.add(id);
    const pars=EDGES.filter(e=>e.to===id).map(e=>e.from);
    if(!pars.length){{depth[id]=0;return 0;}}
    depth[id]=1+Math.max(...pars.map(p=>getDepth(p,vis)));
    return depth[id];
  }}
  NODES.forEach(n=>getDepth(n.id,new Set()));
  // Group by depth
  const levels={{}};
  NODES.forEach(n=>{{const d=depth[n.id]||0;(levels[d]||(levels[d]=[])).push(n.id);}});
  // Place nodes
  const maxLvl=Math.max(...Object.keys(levels).map(Number));
  Object.entries(levels).forEach(([lvl,ids])=>{{
    const w=ids.length*(BW+24)-24;
    ids.forEach((id,i)=>{{
      pos[id]={{x:i*(BW+24)+BW/2-w/2,y:parseInt(lvl)*(BH+90)+80,vx:0,vy:0}};
    }});
  }});
  // Center
  const xs=Object.values(pos).map(p=>p.x);
  if(!xs.length)return;
  const mid=(Math.min(...xs)+Math.max(...xs))/2;
  const cw=cv.width/2/scale;
  Object.values(pos).forEach(p=>p.x+=cw-mid);
  panX=0;panY=60;scale=Math.min(1,cv.width/(Object.keys(pos).length*(BW+24)));
}}
layout();

function simulate(){{
  if(!sim)return;
  const ids=Object.keys(pos);
  for(let i=0;i<ids.length;i++)for(let j=i+1;j<ids.length;j++){{
    const a=pos[ids[i]],b=pos[ids[j]];
    const dx=b.x-a.x,dy=b.y-a.y,d=Math.sqrt(dx*dx+dy*dy)||1;
    const f=7000/(d*d);
    a.vx-=dx/d*f;a.vy-=dy/d*f;b.vx+=dx/d*f;b.vy+=dy/d*f;
  }}
  EDGES.forEach(e=>{{
    const a=pos[e.from],b=pos[e.to];if(!a||!b)return;
    const dx=b.x-a.x,dy=b.y-a.y,d=Math.sqrt(dx*dx+dy*dy)||1;
    const f=(d-180)*0.05;
    a.vx+=dx/d*f;a.vy+=dy/d*f;b.vx-=dx/d*f;b.vy-=dy/d*f;
  }});
  ids.forEach(id=>{{
    if(id!==drag){{pos[id].x+=pos[id].vx;pos[id].y+=pos[id].vy;}}
    pos[id].vx*=0.75;pos[id].vy*=0.75;
  }});
}}

function toW(sx,sy){{return{{x:(sx-panX)/scale,y:(sy-panY)/scale}};}}
function nodeAt(wx,wy){{
  for(let i=NODES.length-1;i>=0;i--){{
    const n=NODES[i],p=pos[n.id];if(!p)continue;
    if(wx>=p.x-BW/2&&wx<=p.x+BW/2&&wy>=p.y-BH/2&&wy<=p.y+BH/2)return n;
  }}return null;
}}
function isVis(nid){{
  let cur=nid,seen=new Set();
  while(true){{
    const pars=EDGES.filter(e=>e.to===cur).map(e=>e.from);
    if(!pars.length)return true;
    for(const p of pars){{if(collapsed.has(p))return false;}}
    if(seen.has(cur))return true;
    seen.add(cur);cur=pars[0];
  }}
}}
function rrect(x,y,w,h,r){{
  ctx.beginPath();ctx.moveTo(x+r,y);ctx.lineTo(x+w-r,y);ctx.arcTo(x+w,y,x+w,y+r,r);
  ctx.lineTo(x+w,y+h-r);ctx.arcTo(x+w,y+h,x+w-r,y+h,r);
  ctx.lineTo(x+r,y+h);ctx.arcTo(x,y+h,x,y+h-r,r);ctx.lineTo(x,y+r);ctx.arcTo(x,y,x+r,y,r);
  ctx.closePath();
}}

function draw(){{
  ctx.clearRect(0,0,cv.width,cv.height);
  ctx.save();ctx.translate(panX,panY);ctx.scale(scale,scale);

  // Draw edges first
  EDGES.forEach(e=>{{
    if(!pos[e.from]||!pos[e.to])return;
    if(!isVis(e.from)||!isVis(e.to))return;
    const a=pos[e.from],b=pos[e.to];
    const isAnd=e.gate==='AND';
    const gc=isAnd?'#7e22ce':'#0369a1';
    ctx.save();
    ctx.beginPath();
    ctx.moveTo(a.x,a.y+BH/2);
    ctx.bezierCurveTo(a.x,a.y+BH/2+35,b.x,b.y-BH/2-35,b.x,b.y-BH/2);
    ctx.strokeStyle=gc;ctx.lineWidth=1.3;ctx.globalAlpha=0.55;ctx.stroke();
    // Gate symbol at midpoint
    const mx=(a.x+b.x)/2,my=(a.y+BH/2+b.y-BH/2)/2;
    ctx.globalAlpha=1;
    ctx.beginPath();ctx.arc(mx,my,GR,0,Math.PI*2);
    ctx.fillStyle=isAnd?'#130018':'#00101e';ctx.fill();
    ctx.strokeStyle=gc;ctx.lineWidth=1;ctx.stroke();
    ctx.fillStyle=isAnd?'#e879f9':'#38bdf8';
    ctx.font='bold 7px JetBrains Mono,monospace';
    ctx.textAlign='center';ctx.textBaseline='middle';
    ctx.fillText(e.gate,mx,my);
    ctx.restore();
  }});

  // Draw nodes
  NODES.forEach(n=>{{
    if(!pos[n.id]||!isVis(n.id))return;
    const p=pos[n.id],x=p.x-BW/2,y=p.y-BH/2;
    const sc=SCOL[n.status]||SCOL.na;
    const tf=TFIL[n.type]||'#0c1628';
    const ts=TSTR[n.type]||'#7ab8e8';
    const hl=searchHL.size===0||searchHL.has(n.id);

    ctx.save();ctx.globalAlpha=hl?1:0.15;
    if(n.status!=='na'&&hl){{ctx.shadowColor=sc;ctx.shadowBlur=12;}}

    // Box
    ctx.fillStyle=tf;rrect(x,y,BW,BH,7);ctx.fill();

    // Multi-parent highlight (cyan dashed)
    if(n.multi_parent&&hl){{
      ctx.strokeStyle='#38bdf8';ctx.lineWidth=2;
      ctx.setLineDash([4,3]);rrect(x-3,y-3,BW+6,BH+6,10);ctx.stroke();ctx.setLineDash([]);
    }}
    // Shared failure highlight (amber dashed)
    if(n.shared&&!n.multi_parent&&hl){{
      ctx.strokeStyle='#f59e0b';ctx.lineWidth=1.5;
      ctx.setLineDash([3,3]);rrect(x-2,y-2,BW+4,BH+4,9);ctx.stroke();ctx.setLineDash([]);
    }}

    // Border
    ctx.shadowBlur=0;
    ctx.strokeStyle=sc;ctx.lineWidth=hl?2:1;
    rrect(x,y,BW,BH,7);ctx.stroke();

    // Type stripe
    ctx.fillStyle=ts+'22';ctx.fillRect(x,y,BW,16);

    // Type label
    ctx.fillStyle=ts;ctx.font='bold 7px JetBrains Mono,monospace';
    ctx.textAlign='center';ctx.textBaseline='top';
    ctx.fillText(n.type+(n.gate&&n.gate!=='–'?' · '+n.gate:''),p.x,y+2);

    // Label
    ctx.fillStyle=ts;ctx.font='bold 11px JetBrains Mono,monospace';
    ctx.textAlign='center';ctx.textBaseline='middle';
    ctx.fillText(n.label.substring(0,16),p.x,p.y-6);

    // Name
    ctx.fillStyle='#4a6a8a';ctx.font='8px DM Sans,sans-serif';
    const nm=n.name.length>22?n.name.substring(0,21)+'…':n.name;
    ctx.fillText(nm,p.x,p.y+6);

    // T and A
    ctx.font='7px JetBrains Mono,monospace';
    ctx.textAlign='left';ctx.fillStyle='#3a7abf';
    ctx.fillText('T:'+n.T,x+5,y+BH-10);
    ctx.textAlign='right';
    ctx.fillStyle=n.status==='pass'?'#22c55e':n.status==='fail'?'#f87171':'#2a4a6a';
    ctx.fillText('A:'+n.A,x+BW-5,y+BH-10);

    // Collapse toggle
    const hasKids=EDGES.some(e=>e.from===n.id);
    if(hasKids){{
      const bx=p.x+BW/2-10,by=y+6;
      ctx.fillStyle=collapsed.has(n.id)?ts+'44':'#1e2d45';
      ctx.beginPath();ctx.arc(bx,by,6,0,Math.PI*2);ctx.fill();
      ctx.strokeStyle=ts;ctx.lineWidth=.8;ctx.stroke();
      ctx.fillStyle=ts;ctx.font='bold 7px monospace';
      ctx.textAlign='center';ctx.textBaseline='middle';
      ctx.fillText(collapsed.has(n.id)?'▶':'▼',bx,by);
    }}
    ctx.restore();
  }});
  ctx.restore();
}}

// Events
cv.addEventListener('mousedown',ev=>{{
  const r=cv.getBoundingClientRect();
  const{{x:wx,y:wy}}=toW(ev.clientX-r.left,ev.clientY-r.top);
  const n=nodeAt(wx,wy);
  if(n&&isVis(n.id)){{
    const p=pos[n.id];
    const bx=p.x+BW/2-10,by=p.y-BH/2+6;
    if(EDGES.some(e=>e.from===n.id)&&Math.hypot(wx-bx,wy-by)<8){{
      collapsed.has(n.id)?collapsed.delete(n.id):collapsed.add(n.id);return;
    }}
    drag=n.id;dragOff={{x:wx-p.x,y:wy-p.y}};cv.classList.add('pan');
  }}else{{isPan=true;lastP={{x:ev.clientX,y:ev.clientY}};cv.classList.add('pan');}}
  ev.preventDefault();
}});
window.addEventListener('mousemove',ev=>{{
  if(drag){{
    const r=cv.getBoundingClientRect();
    const{{x:wx,y:wy}}=toW(ev.clientX-r.left,ev.clientY-r.top);
    pos[drag].x=wx-dragOff.x;pos[drag].y=wy-dragOff.y;
  }}else if(isPan){{panX+=ev.clientX-lastP.x;panY+=ev.clientY-lastP.y;lastP={{x:ev.clientX,y:ev.clientY}};}}
  const r=cv.getBoundingClientRect();
  const{{x:wx2,y:wy2}}=toW(ev.clientX-r.left,ev.clientY-r.top);
  const hn=nodeAt(wx2,wy2);
  const tip=document.getElementById('tip');
  if(hn&&isVis(hn.id)){{
    const sc=hn.status==='pass'?'#22c55e':hn.status==='fail'?'#f87171':'#7ab8e8';
    const multi=hn.multi_parent?'<br><span style="color:#38bdf8">↔ Multi-parent node (DAG) — T = minimum budget across all paths</span>':'';
    const shared=hn.shared?'<br><span style="color:#fbbf24">⚡ Shared failure — achieved value synced across all instances</span>':'';
    tip.innerHTML=`<b style="color:${{sc}}">${{hn.label}}</b> <span style="color:#3a5a7a">${{hn.type}} · gate=${{hn.gate}}</span><br>
      <span style="color:#3a5a7a;font-size:9px">${{hn.name}}</span><br>
      <b style="color:#5aabff">T=${{hn.T}}</b>  <b style="color:${{sc}}">A=${{hn.A}}</b>${{multi}}${{shared}}`;
    tip.style.display='block';
    tip.style.left=Math.min(ev.clientX-r.left+16,cv.width-290)+'px';
    tip.style.top=Math.max(ev.clientY-r.top-70,0)+'px';
  }}else tip.style.display='none';
}});
window.addEventListener('mouseup',()=>{{drag=null;isPan=false;cv.classList.remove('pan');}});
cv.addEventListener('wheel',ev=>{{
  ev.preventDefault();
  const r=cv.getBoundingClientRect();
  const cx=ev.clientX-r.left,cy=ev.clientY-r.top;
  const d=ev.deltaY<0?1.12:0.89;
  const ns=Math.max(0.04,Math.min(8,scale*d));
  panX=cx-(cx-panX)*(ns/scale);panY=cy-(cy-panY)*(ns/scale);scale=ns;
}},{{passive:false}});

function onSearch(q){{
  searchHL.clear();if(!q)return;
  const ql=q.toLowerCase();
  NODES.forEach(n=>{{
    if(n.label.toLowerCase().includes(ql)||n.name.toLowerCase().includes(ql)||n.type.toLowerCase().includes(ql))
      searchHL.add(n.id);
  }});
  const base=new Set(searchHL);
  base.forEach(id=>{{
    let cur=id,seen=new Set();
    while(true){{
      const pars=EDGES.filter(e=>e.to===cur).map(e=>e.from);
      if(!pars.length||seen.has(cur))break;
      seen.add(cur);pars.forEach(p=>searchHL.add(p));cur=pars[0];
    }}
  }});
}}
function zoomIn(){{scale=Math.min(8,scale*1.2);}}
function zoomOut(){{scale=Math.max(0.04,scale/1.2);}}
function resetView(){{scale=1;panX=0;panY=60;searchHL.clear();layout();}}
function toggleSim(){{sim=!sim;document.getElementById('simBtn').textContent='⟳ Physics '+(sim?'ON':'OFF');}}

function loop(){{simulate();draw();requestAnimationFrame(loop);}}
loop();
</script></body></html>"""


# ══════════════════════════════════════════════════════════════════
# SIDEBAR — Node builder (DAG-aware)
# ══════════════════════════════════════════════════════════════════
def sci_input(label, key_m, key_e, default_val=None):
    if default_val and default_val > 0:
        exp = int(math.floor(math.log10(default_val)))
        man = round(default_val / (10 ** exp), 3)
    else:
        exp, man = -7, 1.0
    c1, c2 = st.columns([3,2])
    m = c1.number_input(f"{label} ×", value=man, min_value=0.0, max_value=9.999,
                        step=0.001, format="%.3f", key=key_m)
    e = c2.number_input("10^", value=exp, min_value=-20, max_value=0, step=1, key=key_e)
    return m * (10 ** e) if m > 0 else None

with st.sidebar:
    nodes      = st.session_state.nodes
    edges      = st.session_state.edges
    hz_targets = st.session_state.hz_targets

    ls = st.session_state.get("_last_saved","")
    st.markdown(f'<div class="{"save-ok" if ls else "save-no"}">{"✓ Auto-saved · "+ls if ls else "○ Not saved yet"}</div>',
                unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### ⚛ FTA Builder v11")
    st.caption("DAG mode — nodes can have multiple parents")

    # ── Add Hazard ──────────────────────────────────────────────
    with st.expander("① Add Hazard (HZ)", expanded=False):
        hz_lbl  = st.text_input("Label", value="HZ01", key="hz_lbl")
        hz_nm   = st.text_input("Name",  value="", key="hz_nm")
        hz_gate = st.selectbox("Gate (how HZ splits to its SF children)", ["OR","AND"], key="hz_gate_sel")
        hz_tgt  = sci_input("Target", "hz_m", "hz_e", 1e-7)
        if st.button("➕ Add Hazard", use_container_width=True):
            nid = next_id()
            nodes[nid] = {"id": nid, "label": hz_lbl, "name": hz_nm,
                          "type":"HZ","gate":hz_gate,"achieved":None,"desc":"","parents":[]}
            hz_targets[nid] = hz_tgt or 1e-7
            save_to_file(); st.rerun()

    # ── Add node (supports multiple parents → DAG) ───────────────
    if nodes:
        with st.expander("② Add Node", expanded=False):
            ntype = st.selectbox("Node type", ["SF","FF","IF","AND"], key="ntype")
            valid_types = VALID_PARENT_TYPES.get(ntype,[])
            parent_opts = {k: f"{v.get('label',k)} [{v['type']}]"
                           for k,v in nodes.items() if v["type"] in valid_types}
            if parent_opts:
                n_lbl  = st.text_input("Label", key="n_lbl")
                n_nm   = st.text_input("Name",  key="n_nm")
                n_desc = st.text_input("Desc",  key="n_dsc")
                n_gate = st.selectbox("Gate (how this node splits to its children)",
                                      ["OR","AND"] if ntype not in ("IF",) else ["–"],
                                      key="n_gate")
                st.markdown("**Select parent(s)** — pick one or more to create DAG connections:")
                selected_pars = []
                for pk, pv in list(parent_opts.items())[:30]:  # limit display
                    if st.checkbox(pv, key=f"par_{pk}"):
                        selected_pars.append(pk)
                if st.button(f"➕ Add {ntype}", use_container_width=True):
                    if not n_lbl.strip():
                        st.error("Label required")
                    elif not selected_pars:
                        st.error("Select at least one parent")
                    else:
                        # Check if node with this label already exists → reuse (shared event)
                        existing = nodes_with_label(nodes, n_lbl.strip())
                        if existing:
                            # Add edges from new parents to existing node
                            nid = existing[0]
                            for par in selected_pars:
                                # avoid duplicate edge
                                if not any(e["from"]==par and e["to"]==nid for e in edges):
                                    edges.append({"from":par,"to":nid})
                                    nodes[nid]["parents"].append(par)
                            st.success(f"Linked existing node {n_lbl} to new parent(s)")
                        else:
                            nid = next_id()
                            nodes[nid] = {"id":nid,"label":n_lbl.strip(),"name":n_nm,
                                          "desc":n_desc,"type":ntype,"gate":n_gate,
                                          "achieved":None,"parents":list(selected_pars)}
                            for par in selected_pars:
                                edges.append({"from":par,"to":nid})
                        save_to_file(); st.rerun()
            else:
                st.caption(f"Add {'/'.join(valid_types)} nodes first.")

    # ── Edit node ─────────────────────────────────────────────
    if nodes:
        with st.expander("✏️ Edit node", expanded=False):
            eopts = {k: f"{v.get('label',k)} [{v['type']}]" for k,v in nodes.items()}
            ek = st.selectbox("Select", list(eopts.keys()), format_func=lambda k:eopts[k], key="ek")
            if ek and ek in nodes:
                en = nodes[ek]
                el    = st.text_input("Label", value=en.get("label",""), key="el")
                ename = st.text_input("Name",  value=en.get("name",""),  key="ename")
                edesc = st.text_input("Desc",  value=en.get("desc",""),  key="edesc")
                if en["type"] == "HZ":
                    cur_tgt = hz_targets.get(ek,1e-7)
                    new_tgt = sci_input("HZ Target","etm","ete",cur_tgt)
                if en["type"] not in ("HZ","IF"):
                    egos = ["OR","AND"]
                    eg = st.selectbox("Gate", egos,
                                      index=egos.index(en.get("gate","OR")) if en.get("gate","OR") in egos else 0,
                                      key="eg")
                else:
                    eg = en.get("gate","–")
                if st.button("💾 Save edit", use_container_width=True):
                    nodes[ek].update({"label":el,"name":ename,"desc":edesc,"gate":eg})
                    if en["type"]=="HZ": hz_targets[ek] = new_tgt or 1e-7
                    save_to_file(); st.success("Saved"); st.rerun()

    # ── Delete node ───────────────────────────────────────────
    if nodes:
        with st.expander("🗑 Delete node", expanded=False):
            dopts = {k: f"{v.get('label',k)} [{v['type']}]" for k,v in nodes.items()}
            dk = st.selectbox("Node to delete", list(dopts.keys()),
                              format_func=lambda k:dopts[k], key="dk")
            del_mode = st.radio("Delete mode", ["This node + all descendants","This node only (re-link edges)"], key="del_mode")
            if st.button("🗑 Delete", use_container_width=True):
                if del_mode.startswith("This node +"):
                    # BFS to find all descendants
                    to_del, q = {dk}, [dk]
                    while q:
                        cur = q.pop()
                        for e in edges:
                            if e["from"]==cur and e["to"] not in to_del:
                                to_del.add(e["to"]); q.append(e["to"])
                    for d in to_del:
                        nodes.pop(d,None); hz_targets.pop(d,None)
                    edges[:] = [e for e in edges if e["from"] not in to_del and e["to"] not in to_del]
                else:
                    edges[:] = [e for e in edges if e["from"]!=dk and e["to"]!=dk]
                    nodes.pop(dk,None); hz_targets.pop(dk,None)
                save_to_file(); st.rerun()

    st.markdown("---")

    # ── File ops ─────────────────────────────────────────────
    with st.expander("💾 Save / Load", expanded=False):
        st.markdown("**Download JSON**")
        if nodes:
            st.download_button("⬇ Download JSON", data=json.dumps(state_to_dict(),indent=2),
                               file_name="fta_project_v11.json", mime="application/json",
                               use_container_width=True)
        st.markdown("**Upload JSON**")
        up = st.file_uploader("Upload", type=["json"], key="up_json", label_visibility="collapsed")
        if up:
            try:
                d = json.load(up); dict_to_state(d); save_to_file()
                st.success("✓ Loaded!"); st.rerun()
            except Exception as ex:
                st.error(f"Load failed: {ex}")

        st.markdown("---")
        st.markdown("**Load from GitHub**")
        st.caption("Paste a GitHub file URL (raw or standard). Supports `.json` project files or `.py` app files.")
        gh_url = st.text_input(
            "GitHub URL",
            placeholder="https://github.com/user/repo/blob/main/fta_project.json",
            key="gh_url",
            label_visibility="collapsed",
        )
        if st.button("⬇ Load from GitHub", use_container_width=True, key="gh_load_btn"):
            if not gh_url.strip():
                st.error("Please enter a GitHub URL.")
            else:
                import urllib.request, urllib.error
                # Convert standard GitHub blob URL to raw URL automatically
                raw_url = gh_url.strip()
                if "github.com" in raw_url and "/blob/" in raw_url:
                    raw_url = raw_url.replace("github.com", "raw.githubusercontent.com").replace("/blob/", "/")
                try:
                    with urllib.request.urlopen(raw_url, timeout=10) as resp:
                        content = resp.read().decode("utf-8")
                    if raw_url.endswith(".py"):
                        # For .py files: extract the build_default_data equivalent by exec-ing
                        # the module in isolation and grabbing session state if it sets it,
                        # otherwise show the source for reference
                        st.info("Python file loaded. To use a .py source as data, export your tree as JSON from the app first, then load the JSON here.")
                        st.code(content[:3000] + ("\n... (truncated)" if len(content) > 3000 else ""), language="python")
                    else:
                        # Assume JSON project file
                        d = json.loads(content)
                        dict_to_state(d)
                        save_to_file()
                        st.success(f"✓ Loaded from GitHub! Saved at: {d.get('saved_at', 'unknown')}")
                        st.rerun()
                except urllib.error.HTTPError as e:
                    st.error(f"HTTP error {e.code}: Could not fetch URL. Check the URL is correct and the file is public.")
                except urllib.error.URLError as e:
                    st.error(f"Network error: {e.reason}")
                except json.JSONDecodeError as e:
                    st.error(f"Invalid JSON in file: {e}")
                except Exception as e:
                    st.error(f"Load failed: {e}")

    st.markdown("---")
    if st.button("⚠ Load default HZ04 data", use_container_width=True):
        nd, ed, ht, nv = build_default_data()
        st.session_state.nodes      = nd
        st.session_state.edges      = ed
        st.session_state.hz_targets = ht
        st.session_state.nxt        = nv
        save_to_file(); st.rerun()

    if st.button("⚠ Reset everything", use_container_width=True):
        st.session_state.nodes = {}; st.session_state.edges = []
        st.session_state.hz_targets = {}; st.session_state.nxt = 1
        if os.path.exists(SAVE_FILE): os.remove(SAVE_FILE)
        st.rerun()


# ══════════════════════════════════════════════════════════════════
# MAIN AREA
# ══════════════════════════════════════════════════════════════════
nodes      = st.session_state.nodes
edges      = st.session_state.edges
hz_targets = st.session_state.hz_targets
order      = all_node_ids_reachable(nodes, edges)

alloc  = allocate(nodes, edges, hz_targets)
rolled = rollup(nodes, edges)

# Header
st.markdown("""<div class="app-header">
  <h1>⚛ FTA Risk Allocator v11</h1>
  <p>DAG engine · Multi-parent shared events · OR & AND gate logic · Auto-save · HZ04 Pressurized Fluid Hazard preloaded</p>
</div>""", unsafe_allow_html=True)

# Stats bar
hz_list = [n for n in nodes.values() if n["type"]=="HZ"]
n_sf = sum(1 for v in nodes.values() if v["type"]=="SF")
n_ff = sum(1 for v in nodes.values() if v["type"] in ("FF","AND"))
n_if = sum(1 for v in nodes.values() if v["type"]=="IF")
n_multi = sum(1 for nid in nodes if len(get_parents(edges,nid))>1)

if nodes:
    cols = st.columns(6)
    def mc(l,v,c): return f'<div style="background:#0c1220;border:1px solid #1e2d45;border-radius:6px;padding:8px 14px"><div style="font-size:0.58rem;color:#3a5a7a;text-transform:uppercase;letter-spacing:1px;margin-bottom:2px">{l}</div><div style="font-family:JetBrains Mono,monospace;font-size:1rem;font-weight:700;color:{c}">{v}</div></div>'
    cols[0].markdown(mc("Hazards",len(hz_list),"#fb923c"),unsafe_allow_html=True)
    cols[1].markdown(mc("Sys Failures",n_sf,"#60a5fa"),unsafe_allow_html=True)
    cols[2].markdown(mc("FF / AND",n_ff,"#34d399"),unsafe_allow_html=True)
    cols[3].markdown(mc("Init Events",n_if,"#c084fc"),unsafe_allow_html=True)
    cols[4].markdown(mc("Multi-parent",n_multi,"#38bdf8"),unsafe_allow_html=True)
    cols[5].markdown(mc("Total nodes",len(nodes),"#7ab8e8"),unsafe_allow_html=True)
    st.markdown("<div style='height:8px'></div>",unsafe_allow_html=True)

if not nodes:
    st.markdown('<div class="callout">👈 No data. Use the sidebar to add nodes, or click <b>Load default HZ04 data</b>.</div>',unsafe_allow_html=True)
else:
    tab_tree, tab_vals, tab_table, tab_export = st.tabs(["🌳 Tree","✏️ Values","📋 Table","📥 Export"])

    # ── TAB 1: TREE ───────────────────────────────────────────
    with tab_tree:
        if st.session_state.get("loaded_default"):
            st.markdown('<div class="callout ok">✅ HZ04 Pressurized Fluid Hazard loaded from your FT50–FT61 diagrams. SF-05a and SF-05b are multi-parent shared nodes (cyan border). Hover any node for details.</div>',unsafe_allow_html=True)
        components.html(build_canvas(nodes, edges, alloc, rolled), height=700, scrolling=False)

    # ── TAB 2: VALUES ─────────────────────────────────────────
    with tab_vals:
        st.markdown('<div class="callout">'
            '<b>Allocated (T)</b> = top-down from HZ target. Gate is on the parent node. '
            'Multi-parent nodes receive budget from ALL paths → effective T = minimum (most conservative).<br>'
            '<b>Achieved (A)</b> = value you enter, rolled up bottom-up. '
            '<b>Shared failures</b> (same label) auto-sync to worst-case. '
            '<span style="color:#38bdf8">Cyan border = multi-parent DAG node.</span>'
            '</div>', unsafe_allow_html=True)

        sync_log = []

        for hz in hz_list:
            hid = hz["id"]
            hz_rolled = rolled.get(hid)
            hz_tgt    = hz_targets.get(hid,1e-7)
            ok  = hz_rolled is not None and hz_rolled <= hz_tgt
            bad = hz_rolled is not None and hz_rolled > hz_tgt
            cls = "ok" if ok else ("warn" if bad else "")
            st.markdown(
                f'<div class="callout {cls}">{"✅" if ok else "❌" if bad else "⬜"} '
                f'<b>{hz.get("label","?")} — {hz.get("name","")}</b>'
                f'&emsp; T={fmt(hz_tgt)} &emsp; A={fmt(hz_rolled)}</div>',
                unsafe_allow_html=True)

            sub_ids = [i for i in order if i != hid and i in nodes]

            hc = st.columns([0.8,1.6,2.2,0.5,1.6,2.2,1.6])
            for h,t in zip(hc,["Depth","Label","Name","Type","Allocated (T)","Achieved (enter)","Rolled-up (A)"]):
                h.markdown(f"<span style='font-size:0.58rem;color:#3a5a7a;text-transform:uppercase;letter-spacing:1px'>{t}</span>",unsafe_allow_html=True)

            changed_nid = changed_val = None

            for nid in sub_ids:
                if nid not in nodes: continue
                n   = nodes[nid]
                t   = n["type"]
                alc = alloc.get(nid)
                ach = n.get("achieved")
                rol = rolled.get(nid)
                d   = depth_in_dag(nodes, edges, nid)
                peers      = nodes_with_label(nodes, n.get("label",""))
                is_shared  = len(peers) > 1
                is_multi   = len(get_parents(edges, nid)) > 1
                indent     = "&ensp;" * (d * 2)

                cols = st.columns([0.8,1.6,2.2,0.5,1.6,2.2,1.6])
                cols[0].markdown(f"<span style='color:#2a4a6a;font-size:0.7rem'>{'└─'*min(d,4)}</span>",unsafe_allow_html=True)

                tags = ""
                if is_multi:  tags += '<span class="st-multi">↔ multi-parent</span>'
                if is_shared: tags += '<span class="st-shared">⚡shared</span>'
                tc = TYPE_STR.get(t,"#7ab8e8")
                cols[1].markdown(f"{indent}<span style='font-family:JetBrains Mono,monospace;font-size:0.8rem;color:{tc}'>{n.get('label',nid)}</span>{tags}",unsafe_allow_html=True)
                cols[2].markdown(f"<span style='font-size:0.75rem;color:#4a6a8a'>{n.get('name','')}</span>",unsafe_allow_html=True)
                cols[3].markdown(f"<span class='nb nb-{t}'>{t}</span>",unsafe_allow_html=True)

                # Allocated — show with note if multi-parent
                alc_note = " (min)" if is_multi else ""
                cols[4].markdown(f"<span class='vbadge vb-alloc'>{fmt(alc)}{alc_note}</span>",unsafe_allow_html=True)

                with cols[5]:
                    ca,cb,cc = st.columns([2,1.5,0.8])
                    m_val = round(ach/(10**int(math.floor(math.log10(ach)))),3) if (ach and ach>0) else 1.0
                    e_val = int(math.floor(math.log10(ach))) if (ach and ach>0) else -3
                    new_m = ca.number_input("m",value=m_val,min_value=0.0,max_value=9.999,
                                            step=0.001,format="%.3f",key=f"vm_{nid}",label_visibility="collapsed")
                    new_e = cb.number_input("e",value=e_val,min_value=-20,max_value=0,
                                            step=1,key=f"ve_{nid}",label_visibility="collapsed")
                    if cc.button("✕",key=f"vc_{nid}",help="Clear"):
                        nodes[nid]["achieved"]=None; save_to_file(); st.rerun()
                    else:
                        nv2 = new_m*(10**new_e) if new_m>0 else None
                        if nv2 != ach:
                            changed_nid=nid; changed_val=nv2

                badge_cls = ("vb-ach" + (" over" if (rol and alc and rol>alc) else "")) if rol is not None else "vb-none"
                cols[6].markdown(f"<span class='vbadge {badge_cls}'>{fmt(rol)}</span>",unsafe_allow_html=True)

            if changed_nid and changed_val is not None:
                log = sync_shared(nodes, changed_nid, changed_val)
                sync_log.extend(log); save_to_file(); st.rerun()

            st.markdown("<hr>",unsafe_allow_html=True)

        if sync_log:
            with st.expander(f"🔄 Shared sync — {len(sync_log)} node(s) updated", expanded=True):
                for nid,old,new in sync_log:
                    st.markdown(f"`{nodes.get(nid,{}).get('label',nid)}`: {fmt(old)} → {fmt(new)}")

    # ── TAB 3: TABLE ─────────────────────────────────────────
    with tab_table:
        srch = st.text_input("🔍 Filter", placeholder="label / name / type…", key="tbl_srch")
        rows = []
        for nid in order:
            if nid not in nodes: continue
            n   = nodes[nid]
            alc = alloc.get(nid)
            rol = rolled.get(nid)
            pars = get_parents(edges, nid)
            par_labels = ", ".join(nodes[p].get("label",p) for p in pars if p in nodes) or "–"
            if srch:
                q = srch.lower()
                if not any(q in nodes[nid].get(f,"").lower() for f in ("label","name","type")):
                    continue
            peers = nodes_with_label(nodes, n.get("label",""))
            rows.append({
                "Type":n["type"],"Label":n.get("label",nid),"Name":n.get("name",""),
                "Parent(s)":par_labels,"Gate":n.get("gate","–"),
                "Allocated T":fmt(alc),"Achieved A":fmt(rol),
                "OK?":("✅" if (alc and rol and rol<=alc) else "❌" if (alc and rol and rol>alc) else "–"),
                "Multi-parent":"↔" if len(pars)>1 else "",
                "Shared":"⚡" if len(peers)>1 else "",
            })
        if rows:
            import pandas as pd
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        else:
            st.info("No nodes match." if srch else "No nodes yet.")

    # ── TAB 4: EXPORT ────────────────────────────────────────
    with tab_export:
        c1,c2 = st.columns(2)
        with c1:
            st.markdown("**JSON (full project)**")
            st.download_button("⬇ Download JSON", data=json.dumps(state_to_dict(),indent=2),
                               file_name="fta_v11.json", mime="application/json",
                               use_container_width=True)
        with c2:
            st.markdown("**Excel (.xlsx)**")
            def build_xlsx():
                wb = Workbook(); ws = wb.active; ws.title = "FTA_v11"
                hdrs = ["Type","Label","Name","Desc","Parent(s)","Gate","Allocated T","Achieved A","OK?","Multi-parent","Shared"]
                hf = PatternFill("solid",fgColor="0c1e38"); hfont = Font(name="Consolas",bold=True,color="7ab8e8")
                for ci,h in enumerate(hdrs,1):
                    c = ws.cell(1,ci,h); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center")
                for ri,nid in enumerate(order,2):
                    if nid not in nodes: continue
                    n   = nodes[nid]
                    alc = alloc.get(nid); rol = rolled.get(nid)
                    pars = get_parents(edges,nid)
                    pl = ", ".join(nodes[p].get("label",p) for p in pars if p in nodes) or "–"
                    peers = nodes_with_label(nodes,n.get("label",""))
                    ok = "YES" if (alc and rol and rol<=alc) else "NO" if (alc and rol and rol>alc) else "–"
                    row = [n["type"],n.get("label",""),n.get("name",""),n.get("desc",""),
                           pl,n.get("gate","–"),fmt(alc),fmt(rol),ok,
                           "YES" if len(pars)>1 else "NO","YES" if len(peers)>1 else "NO"]
                    for ci,v in enumerate(row,1): ws.cell(ri,ci,v)
                for col in ["B","C","G","H"]: ws.column_dimensions[col].width=18
                buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()
            st.download_button("⬇ Download Excel", data=build_xlsx(),
                               file_name="fta_v11.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
